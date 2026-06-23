import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def crear_excel():
    wb = openpyxl.Workbook()
    
    # ── ESTILOS COMUNES ────────────────────────────────────────────────────────
    font_title = Font(name="Arial", size=14, bold=True, underline="single")
    font_subtitle = Font(name="Arial", size=11, bold=True)
    font_header = Font(name="Arial", size=10, bold=True)
    font_data = Font(name="Arial", size=10)
    font_bold = Font(name="Arial", size=10, bold=True)

    fill_yellow = PatternFill(start_color="FFE599", end_color="FFE599", fill_type="solid") # Gold/Yellowish
    fill_orange = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid") # Light Orange/Yellow
    fill_pink = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")   # Light Pink/Red
    fill_green_sub = PatternFill(start_color="93C47D", end_color="93C47D", fill_type="solid") # Muted Green

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # ── PESTAÑA 1: Validación Demanda (Powerade) ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "Validación Demanda"
    ws1.views.sheetView[0].showGridLines = True

    # Escribir Título
    ws1["A1"] = "Verificación y Validación:"
    ws1["A1"].font = font_title
    ws1["A2"] = "•  Validación de la cantidad de Powerade vendido por día"
    ws1["A2"].font = font_subtitle
    ws1.row_dimensions[1].height = 25
    ws1.row_dimensions[2].height = 20

    start_row_1 = 4

    # Encabezados Nivel 1
    ws1.merge_cells(start_row=start_row_1, start_column=1, end_row=start_row_1, end_column=3)
    ws1.cell(row=start_row_1, column=1, value="Cantidad de Powerade Generado").fill = fill_yellow
    
    ws1.merge_cells(start_row=start_row_1, start_column=4, end_row=start_row_1, end_column=6)
    ws1.cell(row=start_row_1, column=4, value="Cantidad de Powerade de la medición real").fill = fill_orange
    
    ws1.cell(row=start_row_1, column=7, value="").fill = fill_pink

    for col in range(1, 8):
        cell = ws1.cell(row=start_row_1, column=col)
        cell.font = font_header
        cell.alignment = align_center
        cell.border = thin_border
    ws1.row_dimensions[start_row_1].height = 28

    # Encabezados Nivel 2
    headers1_l2 = [
        "Etiquetas de fila", "Cuenta de Cantidad Vendida", "Frecuencia relativa",
        "Etiquetas de fila", "Cuenta de etiquetas de fila", "Frecuencia relativa",
        "Error"
    ]
    for col_idx, text in enumerate(headers1_l2, start=1):
        cell = ws1.cell(row=start_row_1 + 1, column=col_idx, value=text)
        cell.font = font_header
        cell.alignment = align_center
        cell.border = thin_border
        if col_idx == 7:
            cell.fill = fill_pink
        else:
            cell.fill = fill_green_sub
    ws1.row_dimensions[start_row_1 + 1].height = 35

    # Datos
    datos1 = [
        (0, 7237, 44), (1, 10696, 66), (2, 7297, 45), (3, 7561, 46),
        (4, 6247, 39), (5, 4226, 26), (6, 1937, 12), (7, 1456, 9),
        (8, 1209, 7), (9, 1176, 7), (10, 325, 2), (11, 330, 2),
        (12, 145, 1), (13, 158, 1)
    ]

    total_row_1 = start_row_1 + 2 + len(datos1) # Fila 20 es el total
    
    for i, (val, cant_gen, cant_real) in enumerate(datos1):
        curr_row = start_row_1 + 2 + i
        ws1.cell(row=curr_row, column=1, value=val).alignment = align_center
        ws1.cell(row=curr_row, column=2, value=cant_gen).number_format = '#,##0'
        ws1.cell(row=curr_row, column=3, value=f"=B{curr_row}/B{total_row_1}").number_format = '0.00%'
        
        ws1.cell(row=curr_row, column=4, value=val).alignment = align_center
        ws1.cell(row=curr_row, column=5, value=cant_real).number_format = '#,##0'
        ws1.cell(row=curr_row, column=6, value=f"=E{curr_row}/E{total_row_1}").number_format = '0.00%'
        
        ws1.cell(row=curr_row, column=7, value=f"=ABS(C{curr_row}-F{curr_row})").number_format = '0.00%'
        
        for col_idx in range(1, 8):
            cell = ws1.cell(row=curr_row, column=col_idx)
            cell.font = font_data
            cell.border = thin_border
            if col_idx in [2, 5, 3, 6, 7]:
                cell.alignment = align_right
        ws1.row_dimensions[curr_row].height = 20

    # Totales Fila 20
    ws1.cell(row=total_row_1, column=1, value="Total general").fill = fill_pink
    ws1.cell(row=total_row_1, column=1).font = font_bold
    ws1.cell(row=total_row_1, column=1).alignment = align_center
    ws1.cell(row=total_row_1, column=1).border = thin_border

    ws1.cell(row=total_row_1, column=2, value=f"=SUM(B{start_row_1+2}:B{total_row_1-1})").font = font_bold
    ws1.cell(row=total_row_1, column=2).number_format = '#,##0'
    ws1.cell(row=total_row_1, column=2).border = thin_border

    ws1.cell(row=total_row_1, column=3, value="").border = thin_border

    ws1.cell(row=total_row_1, column=4, value="Total general").fill = fill_pink
    ws1.cell(row=total_row_1, column=4).font = font_bold
    ws1.cell(row=total_row_1, column=4).alignment = align_center
    ws1.cell(row=total_row_1, column=4).border = thin_border

    ws1.cell(row=total_row_1, column=5, value=f"=SUM(E{start_row_1+2}:E{total_row_1-1})").font = font_bold
    ws1.cell(row=total_row_1, column=5).number_format = '#,##0'
    ws1.cell(row=total_row_1, column=5).border = thin_border

    ws1.cell(row=total_row_1, column=6, value="").border = thin_border

    # G20 y G21 combinada para promedio de error
    ws1.merge_cells(start_row=total_row_1, start_column=7, end_row=total_row_1+1, end_column=7)
    c_merged_g1 = ws1.cell(row=total_row_1, column=7, value=f"=AVERAGE(G{start_row_1+2}:G{total_row_1-1})")
    c_merged_g1.font = font_bold
    c_merged_g1.fill = fill_pink
    c_merged_g1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c_merged_g1.number_format = '"Error promedio = "\n0.00%'
    c_merged_g1.border = thin_border
    ws1.cell(row=total_row_1+1, column=7).border = thin_border

    ws1.row_dimensions[total_row_1].height = 25
    ws1.row_dimensions[total_row_1+1].height = 25

    # Ajustar anchos
    widths1 = {1: 18, 2: 26, 3: 20, 4: 18, 5: 26, 6: 20, 7: 24}
    for col in range(1, 8):
        ws1.column_dimensions[get_column_letter(col)].width = widths1[col]


    # ── PESTAÑA 2: Validación Demora ──────────────────────────────────────────────
    ws2 = wb.create_sheet(title="Validación Demora")
    ws2.views.sheetView[0].showGridLines = True

    # Escribir Título
    ws2["A1"] = "Verificación y Validación:"
    ws2["A1"].font = font_title
    ws2["A2"] = "•  Validación de la cantidad de días de demora"
    ws2["A2"].font = font_subtitle
    ws2.row_dimensions[1].height = 25
    ws2.row_dimensions[2].height = 20

    start_row_2 = 4

    # Encabezados
    headers2 = [
        "Etiquetas de fila", "Cuenta de Días de Demora", "Frecuencia relativa",
        "Frecuencia esperada", "Error"
    ]
    for col_idx, text in enumerate(headers2, start=1):
        cell = ws2.cell(row=start_row_2, column=col_idx, value=text)
        cell.font = font_header
        cell.alignment = align_center
        cell.border = thin_border
        if col_idx == 5:
            cell.fill = fill_pink
        else:
            cell.fill = fill_green_sub
    ws2.row_dimensions[start_row_2].height = 30

    # Datos
    datos2 = [
        (1, 16535, 1/3),
        (2, 16701, 1/3),
        (3, 16764, 1/3)
    ]

    total_row_2 = start_row_2 + 1 + len(datos2) # Fila 8 es el total

    for i, (val, cuenta, esperada) in enumerate(datos2):
        curr_row = start_row_2 + 1 + i
        ws2.cell(row=curr_row, column=1, value=val).alignment = align_center
        ws2.cell(row=curr_row, column=2, value=cuenta).number_format = '#,##0'
        ws2.cell(row=curr_row, column=3, value=f"=B{curr_row}/B{total_row_2}").number_format = '0.00%'
        ws2.cell(row=curr_row, column=4, value=esperada).number_format = '0.00%'
        ws2.cell(row=curr_row, column=5, value=f"=ABS(C{curr_row}-D{curr_row})").number_format = '0.00%'
        
        for col_idx in range(1, 6):
            cell = ws2.cell(row=curr_row, column=col_idx)
            cell.font = font_data
            cell.border = thin_border
            if col_idx in [2, 3, 4, 5]:
                cell.alignment = align_right
        ws2.row_dimensions[curr_row].height = 22

    # Totales Fila 8
    ws2.cell(row=total_row_2, column=1, value="Total general").fill = fill_pink
    ws2.cell(row=total_row_2, column=1).font = font_bold
    ws2.cell(row=total_row_2, column=1).alignment = align_center
    ws2.cell(row=total_row_2, column=1).border = thin_border

    ws2.cell(row=total_row_2, column=2, value=f"=SUM(B{start_row_2+1}:B{total_row_2-1})").font = font_bold
    ws2.cell(row=total_row_2, column=2).number_format = '#,##0'
    ws2.cell(row=total_row_2, column=2).border = thin_border

    ws2.cell(row=total_row_2, column=3, value="").border = thin_border
    ws2.cell(row=total_row_2, column=4, value="").border = thin_border

    # E8 y E9 combinada para Error Promedio
    ws2.merge_cells(start_row=total_row_2, start_column=5, end_row=total_row_2+1, end_column=5)
    c_merged_g2 = ws2.cell(row=total_row_2, column=5, value=f"=AVERAGE(E{start_row_2+1}:E{total_row_2-1})")
    c_merged_g2.font = font_bold
    c_merged_g2.fill = fill_pink
    c_merged_g2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c_merged_g2.number_format = '"Error Promedio"\n0.00%'
    c_merged_g2.border = thin_border
    ws2.cell(row=total_row_2+1, column=5).border = thin_border

    ws2.row_dimensions[total_row_2].height = 25
    ws2.row_dimensions[total_row_2+1].height = 25

    # Ajustar anchos
    widths2 = {1: 18, 2: 26, 3: 20, 4: 20, 5: 22}
    for col in range(1, 6):
        ws2.column_dimensions[get_column_letter(col)].width = widths2[col]


    # Guardar libro
    nombre_archivo = "Verificacion_y_Validacion_Referencia.xlsx"
    wb.save(nombre_archivo)
    print(f"Archivo '{nombre_archivo}' generado con éxito con 2 pestañas.")

if __name__ == "__main__":
    crear_excel()
